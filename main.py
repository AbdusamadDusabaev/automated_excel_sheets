import datetime
from openpyxl import load_workbook
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive


folder_name = "Av Steel"
file_name_1 = "1.xlsx"
file_name_2 = "2.xlsx"


def add_additional_price(start_price):
    if start_price < 800:
        percent = 0.55
    elif 800 <= start_price < 1200:
        percent = 0.5
    elif 1200 <= start_price < 1500:
        percent = 0.45
    elif 1500 <= start_price < 2000:
        percent = 0.4
    elif 2000 <= start_price < 3000:
        percent = 0.35
    elif 3000 <= start_price < 5000:
        percent = 0.3
    else:
        percent = 0.25
    additional_price = start_price * percent
    final_price = int(start_price + additional_price)
    return final_price


def get_first_data():
    first_data = dict()
    workbook = load_workbook(file_name_1)
    current_page = workbook.active
    for index in range(13, current_page.max_row + 1):
        product_id = current_page[f"J{index}"].value
        start_price = current_page[f"O{index}"].value
        price = add_additional_price(start_price=start_price)
        first_data[product_id] = price
    return first_data


def check_data(first_data):
    first_product_ids = list(first_data.keys())
    workbook = load_workbook(file_name_2)
    current_page = workbook.active

    columns = dict()
    for symbol in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J",
                   "K", "L", "M", "N", "O", "P", "Q",
                   "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA"]:
        if current_page[f"{symbol}1"].value == "Id":
            columns["id"] = symbol
        elif current_page[f"{symbol}1"].value == "Price":
            columns["price"] = symbol
        elif current_page[f"{symbol}1"].value == "DateEnd":
            columns["date_end"] = symbol

    for index in range(2, current_page.max_row + 1):
        product_id = current_page[f"{columns['id']}{index}"].value
        if product_id in first_product_ids:
            current_page[f"{columns['price']}{index}"].value = first_data[product_id]
            current_page[f"{columns['date_end']}{index}"].value = ""
            print(f"[INFO] Товар с артикулом {product_id} найден в обеих таблицах")
        else:
            current_page[f"{columns['price']}{index}"].value = ""
            current_page[f"{columns['date_end']}{index}"].value = str(datetime.date.today())
            print(f"[INFO] Товар с артикулом {product_id} не найден во второй таблице")
    workbook.save(file_name_2)


def update_prices():
    print("[INFO] Обновляем цены в Excel Таблицах")
    print("[INFO] Получаем данные из первой таблицы")
    first_data = get_first_data()
    print("[INFO] Данные из первой таблицы получены")
    print("[INFO] Сравниваем данные первой и второй таблицы")
    check_data(first_data=first_data)
    print("[INFO] Данные обработаны и перезаписаны")


def download_file(google_auth, file_name):
    drive = GoogleDrive(google_auth)
    query = f"title='{file_name}' and trashed=false"
    file = drive.ListFile({'q': query}).GetList()[0]
    file_object = drive.CreateFile({"id": file["id"]})
    file_object.GetContentFile(file["title"])
    print(f"[INFO] Файл {file_name} успешно загружен из Google Диска")
    file_object.Delete()


def download_google_drive(google_auth):
    print("[INFO] Скачиваем документы с Google Диска")
    download_file(file_name=file_name_1, google_auth=google_auth)
    download_file(file_name=file_name_2, google_auth=google_auth)
    print("[INFO] Все документы успешно скачаны")


def upload_google_drive(google_auth):
    drive = GoogleDrive(google_auth)
    query = f"title='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    folder = drive.ListFile({'q': query}).GetList()[0]
    file_1 = drive.CreateFile({'parents': [{'id': folder['id'], 'title': file_name_1}]})
    file_1.SetContentFile(file_name_1)
    file_1.Upload()
    print(f"[INFO] Файл {file_name_1} успешно загружен на Google Диск")
    file_2 = drive.CreateFile({'parents': [{'id': folder['id'], 'title': file_name_2}]})
    file_2.SetContentFile(file_name_2)
    file_2.Upload()
    print(f"[INFO] Файл {file_name_2} успешно загружен на Google Диск")
    print("[INFO] Все данные успешно загружены на Google Диск")


def auth_google():
    google_auth = GoogleAuth()
    google_auth.LocalWebserverAuth()
    return google_auth


def main():
    print("[INFO] Программа запущена")
    print("[INFO] Проходим аутентификацию в Google Диске")
    google_auth = auth_google()
    download_google_drive(google_auth=google_auth)
    update_prices()
    upload_google_drive(google_auth=google_auth)
    print("[INFO] Программа завершена")


if __name__ == "__main__":
    main()
